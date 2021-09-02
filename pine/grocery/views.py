from django.shortcuts import render, redirect
from .models import *
# for profile linking
from accounts.models import Profile
from django.core.exceptions import PermissionDenied
from django.contrib.auth.decorators import login_required, user_passes_test
# adding for contact form
from main.models import Location, City
from django.contrib.auth.models import User
from django.contrib import messages
from django.core.exceptions import ObjectDoesNotExist
from django.http import FileResponse
from .models import Product, Brand, Shop, Category
from .forms import *
import openpyxl as xl
from .email import *
from django.shortcuts import render
from django.http import Http404
from django.http import HttpResponse
import json
from django.conf import settings
import logging
import os
import time
import shutil
from django.template.defaultfilters import slugify

# logging.basicConfig(filename="logging.log", level=logging.WARNING, filemode="a+")
import random


def user_verified(user):
    try:
        return user.profile.is_verified()
    except ObjectDoesNotExist:
        return False


def shop_verified(user):
    try:
        return user.profile.is_shop_owner()
    except ObjectDoesNotExist:
        return False


@login_required
def cart_add(request, shopname, shop_location):
    user_id = request.user.pk
    shop_name = slugify(shopname)
    shop = Shop.objects.get(shop=shop_name, location=shop_location)
    profile = Profile.objects.get(user=user_id)
    product_id = request.POST.get('product_id')
    quantity = 1
    file_path = settings.BASE_DIR + '/media/json/active/user_' + str(profile.id) + '/shop_' + str(shop.id) + '.json'
    try:
        with open(file_path, 'r') as json_read:
            data = json.loads(json_read.read())
        try:
            if (data[str(product_id)] >= 0):
                data[str(product_id)] += quantity
        except:
            data[str(product_id)] = quantity
        with open(file_path, 'w+') as f:
            json.dump(data, f)
    except:
        with open(file_path, 'w+') as json_file:
            add = {}
            quantity = 1
            add[str(product_id)] = quantity
            json.dump(add, json_file)
    return render(request, 'cart.html', {'shop_name': shop_name, 'shop_location': shop_location})


@login_required
def show_savings(cart):
    total_savings = 0
    for product in cart.keys():
        prod_obj = Product.objects.get(id=product)
        total_savings = total_savings + cart[product] * (prod_obj.off)
    return total_savings


@login_required
def updateCart(request, shopname, shop_location):
    shop_name = slugify(shopname)
    shop = Shop.objects.get(shop=shop_name, location=shop_location)
    user_id = request.user.pk
    profile = Profile.objects.get(user=user_id)
    product_id = request.POST.get('product_id')
    quantity = request.POST.get('quantity')
    print(product_id, quantity)
    quantity = int(quantity)
    print(quantity)
    file_path = settings.BASE_DIR + '/media/json/active/user_' + str(profile.id) + '/shop_' + str(shop.id) + '.json'
    with open(file_path, 'r+') as json_read:
        data = json.loads(json_read.read())
    data[str(product_id)] = quantity
    with open(file_path, 'w+') as f:
        json.dump(data, f)
    return render(request, 'cart.html', {'cart': data, 'shop_name': shop_name, 'shop_location': shop_location})


@login_required
def cart_empty(request, shopname):
    user_id = request.user.pk
    shop_name = slugify(shopname)
    shop = Shop.objects.get(shop=shop_name)
    profile = Profile.objects.get(user=user_id)
    email_id = request.POST.get('email')
    file_path = settings.BASE_DIR + '/media/json/active/user_' + str(profile.id) + '/shop_' + str(shop.id) + '.json'
    try:
        with open(file_path, 'w+') as json_new:
            d = {}
            json.dump(d, json_new)
            return render(request, 'page2.html')
    except:
        # raise Http404("Action Cannot be executed!")
        pass


@user_passes_test(shop_verified, login_url='/account/settings/')
def data_upload(request):
    pk = request.user.pk
    user = User.objects.get(pk=pk)
    profile = Profile.objects.get(user=user)
    shop_user = Shop.objects.get(shop_user=profile)

    if request.method == "POST":
        excel_file = request.FILES["excel_file"]
        wb = xl.load_workbook(excel_file)
        sheet = wb.active
        # print(sheet)
        max_r = sheet.max_row
        dic = []
        for i in range(1, max_r + 1):
            if (i != 1):
                name = str(sheet.cell(row=i, column=2).value),
                quantity = str(sheet.cell(row=i, column=3).value),
                price = int(sheet.cell(row=i, column=4).value),
                selling_price = int(sheet.cell(row=i, column=5).value),
                description = str(sheet.cell(row=i, column=6).value),
                # barcode       = str(sheet.cell(row=i,column=7).value),

                off = False
                if price != selling_price:
                    off = True
                    savings = price[0] - selling_price[0]
                    print(savings)
                else:
                    savings = 0
                dic.append(Product(
                    # id            = str(sheet.cell(row=i,column=1).value),
                    name=str(sheet.cell(row=i, column=2).value),
                    quantity=str(sheet.cell(row=i, column=3).value),
                    price=str(sheet.cell(row=i, column=4).value),
                    selling_price=str(sheet.cell(row=i, column=5).value),
                    description=str(sheet.cell(row=i, column=6).value),
                    # barcode       = str(sheet.cell(row=i,column=7).value),
                    shop=shop_user,
                    off=str(off),
                    savings=str(savings),
                ))
                print(shop_user)
        Product.objects.bulk_create(dic)
        return HttpResponse("Data created at server for" + " ")

        Product.objects.bulk_create(dic)
        return HttpResponse("Data created at server for" + " ")

    else:
        return render(request, "upload_data.html")


@user_passes_test(shop_verified, login_url='/account/settings/')
def data_upload_table(request):
    pk = request.user.pk
    user = User.objects.get(pk=pk)
    profile = Profile.objects.get(user=user)
    shop = Shop.objects.get(shop_user=profile)
    groceries = Product.objects.filter(shop=shop)

    return render(request, 'grocery_list_edit.html', {'groceries': groceries, 'form': 'form'})


@user_passes_test(shop_verified, login_url='/account/settings/')
def data_upload_form(request, product_id):
    pk = request.user.pk
    user = User.objects.get(pk=pk)
    profile = Profile.objects.get(user=user)
    shop = Shop.objects.get(shop_user=profile)
    groceries = Product.objects.filter(shop=shop)
    items = Product.objects.get(id=product_id)
    form_class = ProductForm(shop, instance=items)
    images = Images.objects.filter(product_image=product_id).first()
    imageform = ImageForm(instance=images)
    if request.method == 'POST':
        form = ProductForm(shop, request.POST, instance=items)
        # if images is not None :
        #     imageform = ImageForm(request.POST, request.FILES, instance=images)
        # else :
        imageform = ImageForm(request.POST, request.FILES)
        if form.is_valid() and imageform.is_valid():
            post_form = form.save(commit=False)
            post_form.shop = shop
            post_form.instance = items

            post_form.save()
            if images is not None:
                photo = Images(id=images.id, product_image=post_form, image=imageform.instance.image)
            else:
                photo = Images(product_image=post_form, image=imageform.instance.image)
            if imageform.cleaned_data['image']:
                photo.save()
            return redirect('Product_Edit')
        else:
            print(form.errors, imageform.errors)
    else:
        form = form_class
    return render(request, 'grocery_list_edit.html',
                  {'product_id': product_id, 'groceries': groceries, 'form': form, 'imageform': imageform})


def all_shops(request):
    if request.user.is_authenticated and request.method != "POST":
        pk = request.user.pk
        user = User.objects.get(pk=pk)
        profile = Profile.objects.get(user=user)

        if (profile.location):
            location = profile.location
            city = City.objects.get(id=location.city.id)
            shops = Shop.objects.filter(location__city=city)
            form = ShopLocationForm(initial={'city': city})
        else:
            shops = Shop.objects.all()
            for shop in shops:
                print(shop.shop, shop.shop_user, shop.location.city, shop.delivery_at)
            form = ShopLocationForm()

    else:
        if request.method == "POST":
            form = ShopLocationForm(request.POST)
            if form.is_valid():
                city = form.cleaned_data['city']
                print(city)
                shops = Shop.objects.filter(location__city=city)
                for shop in shops:
                    print(shop.shop, shop.shop_user, shop.location.city)

        else:
            shops = Shop.objects.all()
            for shop in shops:
                print(shop.shop, shop.shop_user, shop.location.city)
            form = ShopLocationForm()
    for shop in shops:
        print(shop.shop, shop.shop_user, shop.location.city, shop.delivery_at.all())

    return render(request, 'shops.html', {'shops': shops, 'form': form})


def all_carts(request):
    if request.user.is_authenticated and request.method != "POST":
        pk = request.user.pk
        user = User.objects.get(pk=pk)
        profile = Profile.objects.get(user=user)

        if (profile.location):
            location = profile.location
            city = City.objects.get(id=location.city.id)
            shops = Shop.objects.filter(location__city=city)
            form = ShopLocationForm(initial={'city': city})
        else:
            shops = Shop.objects.all()
            for shop in shops:
                print(shop.shop, shop.shop_user, shop.location.city)
            form = ShopLocationForm()

    else:
        if request.method == "POST":
            form = ShopLocationForm(request.POST)
            if form.is_valid():
                city = form.cleaned_data['city']
                print(city)
                shops = Shop.objects.filter(location__city=city)
                for shop in shops:
                    print(shop.shop, shop.shop_user, shop.location.city)

        else:
            shops = Shop.objects.all()
            for shop in shops:
                print(shop.shop, shop.shop_user, shop.location.city)
            form = ShopLocationForm()
    return render(request, 'all_carts.html', {'shops': shops, 'form': form})


def request_url(request):
    S = request.path
    S = ((S.split('/'))[1].split('/')[0])
    return str(S)


def shops_grocery(request, shopname, shop_location):
    shopname = slugify(shopname)
    shop = Shop.objects.get(shop=shopname, location=shop_location)
    groceries = Product.objects.filter(shop=shop, show_product=True)
    deliverable_locations = shop.delivery_at.all()
    print(deliverable_locations)
    user_id = request.user.pk
    dic = []
    can_be_delivered = None
    profile = None
    try:
        profile = Profile.objects.get(user=user_id)
        profile_location = profile.location
        if profile.location in deliverable_locations:
            can_be_delivered = True
        else:
            can_be_delivered = False
        if request.user.is_authenticated:
            file_path = settings.BASE_DIR + '/media/json/active/user_' + str(profile.id) + '/shop_' + str(
                shop.id) + '.json'
            try:
                with open(file_path, 'r+') as json_cart:
                    user_cart = json.loads(json_cart.read())
                    for key, values in user_cart.items():
                        dic.append(int(key))
            except:
                try:
                    os.mkdir('media/json/active/user_' + str(profile.id))
                except:
                    pass
        else:
            raise PermissionDenied
        print(dic)
    except:
        profile_location = None
        pass

    return render(request, 'groceries.html',
                  {'groceries': groceries, 'shop_name': shop.shop,
                   'shop_location': shop_location, 'deliverable_locations': deliverable_locations,
                   'profile_location': profile_location, 'dic': dic, 'shop': shop,
                   'profile': profile,
                   'can_be_delivered': can_be_delivered})


def shops_by_name(request, shopname):
    shopname = slugify(shopname)
    try:
        shop = Shop.objects.filter(shop=shopname)
        if shop.count() == 1:
            shop = Shop.objects.get(shop=shopname)
            location_id = shop.location.id
            return redirect('Shops_Grocery', shopname, location_id)
    except:
        return redirect('shops')

    if request.method == "POST":
        form = ShopLocationForm(request.POST)
        if form.is_valid():
            city = form.cleaned_data['city']
            shops = Shop.objects.filter(location__city=city)
    else:
        shops = Shop.objects.all()
        form = ShopLocationForm()
    return render(request, 'shops.html', {'shops': shops, 'form': form})


def all_category(request):
    category = Category.objects.all()
    shop_id = request.GET.get('shop', None)
    category_id = request.GET.get('category', None)

    return render(request, '#', {'category': category})


def category_grocery(request, shop_name, category_name):
    shop = Shops.objects.get(shop=shop_name)
    category = Category.objects.filter(category=category_name)
    groceries = Products.objects.filter(shop=shop, category=category)

    return render(request, 'groceries.html', {'groceries': groceries})


def all_brands(request):
    brand = Brand.objects.all()
    groceries = Products.objects.filter(brand=brand)
    return render(request, '#', {'brands': brands})


def all_brands(request, shop_name, brand_name):
    shop = Shops.objects.get(shop=shop_name)
    category = Category.objects.filter(category=category_name)
    groceries = Products.objects.filter(shop=shop, category=category)

    return render(request, 'groceries.html', {'groceries': groceries})


@login_required
def cart_view(request, shopname, shop_location):
    user_id = request.user.pk
    shop_name = slugify(shopname)
    shop = Shop.objects.get(shop=shop_name, location=shop_location)
    profile = Profile.objects.get(user=user_id)
    file_path = settings.BASE_DIR + '/media/json/active/user_' + str(profile.id) + '/shop_' + str(shop.id) + '.json'
    deliverable_locations = shop.delivery_at.all()
    if profile.location in deliverable_locations:
        can_be_delivered = True
    else:
        can_be_delivered = False
    if (request.POST):
        location_id = profile.location.id

        return redirect('Checkout', shop_name, shop_location, location_id)
    try:
        with open(file_path, 'r+') as json_cart:
            user_cart = json.loads(json_cart.read())
            dic = {}
            for key, values in user_cart.items():
                li = []
                pro = Product.objects.filter(id=key).first()
                try:
                    im = Images.objects.filter(product_image=key).first()
                    imn = im.image.url
                except:
                    imn = "/static/assets/img/no-image.png"
                li.append(pro.price)
                li.append(values)
                li.append(pro.name)
                li.append(imn)
                li.append(pro.selling_price)
                li.append(pro.savings)
                dic[key] = li
    except:
        return render(request, 'cart.html',
                      {'can_be_delivered': can_be_delivered, 'cart': {}, 'shop_name': shop_name, 'shop': shop,
                       'shop_location': shop_location, 'profile': profile})
    # form = DeliveryLocationForm()
    return render(request, 'cart.html',
                  {'can_be_delivered': can_be_delivered, 'cart': dic, 'profile': profile, 'shop_name': shop_name,
                   'shop_location': shop_location, 'shop': shop, })


@login_required
def removeItem(request, shopname, shop_location):
    user_id = request.user.pk
    shop_name = slugify(shopname)
    shop = Shop.objects.get(shop=shop_name, location=shop_location)
    profile = Profile.objects.get(user=user_id)
    file_path = settings.BASE_DIR + '/media/json/active/user_' + str(profile.id) + '/shop_' + str(shop.id) + '.json'
    product_id = request.POST.get('product_id')
    print(request.POST.get)
    with open(file_path, 'r') as json_cart:
        user_cart = json.loads(json_cart.read())
    del user_cart[str(product_id)]
    with open(file_path, 'w+') as json_cart:
        json.dump(user_cart, json_cart)
    print(user_cart)
    return render(request, 'cart.html', {'cart': user_cart, 'shop_name': shop_name, 'shop_location': shop_location})


@user_passes_test(user_verified, login_url='/account/settings/')
def checkout(request, shopname, location_id, shop_location):
    user_id = request.user.pk
    shop_name = slugify(shopname)
    shop = Shop.objects.get(shop=shop_name, location=shop_location)
    profile = Profile.objects.get(user=user_id)
    file_path = settings.BASE_DIR + '/media/json/active/user_' + str(profile.id) + '/shop_' + str(shop.id) + '.json'
    dic3 = {}
    dic2 = {}
    prod_dic = {}
    print(request.POST)
    location = Location.objects.get(id=location_id)
    amount = 0
    # try:
    with open(file_path, 'r+') as json_cart:

        dic = {}
        user_cart = json.loads(json_cart.read())
        for key, value in user_cart.items():
            prod_details_dic = {}

            li = []
            pro = Product.objects.filter(id=key).first()
            try:
                im = Images.objects.filter(product_image=key).first()
                imn = im.image.url
            except:
                imn = "/static/assets/img/no-image.png"

            li.append(pro.name)  # 0
            li.append(pro.price)  # 1
            li.append(pro.selling_price)  # 2
            li.append(value)  # 3 ----"Quantity order" ----- #
            li.append(imn)  # 4
            li.append(pro.shop.shop)  # 5
            dic[key] = li  # 6
            prod_details_dic["Product_Name"] = pro.name
            prod_details_dic["MRP"] = pro.price
            prod_details_dic["SP"] = pro.selling_price
            prod_details_dic["Quantity"] = value
            prod_details_dic["delivered"] = ""
            prod_details_dic["image"] = imn
            value = int(value)
            amount += pro.selling_price * value
            id_key = str(key)
            prod_dic[id_key] = prod_details_dic
        shopName = pro.shop.id
    # except:
    #     print("hi")
    #     return Http404

    try:
        user_name = profile.title + profile.name
        # make the name compulsory for every user
        details = []
        details.append(dic)
        details.append(profile.mobile_no)
        details.append(profile.email)
        dic2[user_name] = details
        local = []
        local.append(location.location)
        dic2['locality'] = local
        dic3["name"] = profile.name
        dic3["location"] = location.location
        dic3["mobile_no"] = profile.mobile_no
        dic3["orders"] = prod_dic
        dic3["email"] = profile.email
        dic3['cancelled_by_user'] = "No"
        dic3['cancelled_by_shop'] = "No"
        dic3['delivered'] = "No"
        dic3['amount'] = amount

    except:
        print('hi2')
        pass
    # function to send mail
    file_path2 = 'media/json/shops/shop_' + str(shop.id) + '/' + str(
        time.strftime('%Y-%m-%d--%H-%M-%S')) + '_' + 'user_' + str(profile.id) + '.json'
    file_path3 = 'media/json/orders/user_' + str(profile.id) + "/" + str(
        time.strftime('%Y-%m-%d--%H-%M-%S')) + "_shop_" + str(shop.id) + '.json'

    try:
        os.mkdir('media/json/shops/shop_' + str(shop.id))
    except:
        print('h3')
        pass

    try:
        os.mkdir('media/json/orders/user_' + str(profile.id))
    except:
        print('h3')
        pass

    try:
        with open(file_path2, 'a+') as json_file:
            json.dump(dic2, json_file)
    except:
        print('h4')
        pass

    try:
        with open(file_path3, 'a+') as json_file:
            json.dump(dic3, json_file)
    except:
        print('h5')
        pass

    order = Order.objects.create(cart=file_path3, shop_id=shop.id, user_id=profile.id)
    order.save()
    open(file_path, 'w').close()
    """send_mail_order(profile.id,shopName,dic,dic2)
    send_mail_receipt(profile.id,shopName,dic,dic2)"""

    return redirect('/')
